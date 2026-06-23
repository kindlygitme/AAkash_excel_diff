import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from datetime import datetime
import base64

st.set_page_config(
    page_title="Aakash ExcelDiff Pro",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ── Logo ──────────────────────────────────────────────────────────
LOGO_URL = "https://raw.githubusercontent.com/kindlygitme/aakash-timestamp/0e175e042c4ec15f6c0170e04179ac65526eae55/Aakash%20Logo.png"
import urllib.request, base64
try:
    with urllib.request.urlopen(LOGO_URL) as resp:
        LOGO_B64 = base64.b64encode(resp.read()).decode()
except Exception:
    # Fallback to local upload if available
    try:
        with open("/mnt/user-data/uploads/Aakash_Logo.png", "rb") as f:
            LOGO_B64 = base64.b64encode(f.read()).decode()
    except Exception:
        LOGO_B64 = ""

# ═══════════════════════════════════════════════════════════════════
# CSS – Aakash brand palette: Aakash Blue #009FE3, Red #E52526, White
# ═══════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
.stApp { background: #F0F4F8; color: #1A2B3C; }
#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 0 2.5rem 4rem 2.5rem; max-width: 1400px; }

/* ── Top Nav Bar ── */
.aakash-nav {
    background: #FFFFFF;
    border-bottom: 3px solid #009FE3;
    padding: 0.6rem 2.5rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 0;
    box-shadow: 0 2px 12px rgba(0,159,227,0.10);
}
.aakash-nav-logo { display: flex; align-items: center; gap: 0.7rem; }
.aakash-nav-logo img { height: 48px; object-fit: contain; }
.aakash-nav-title {
    font-size: 1.05rem; font-weight: 700; color: #009FE3;
    letter-spacing: 0.04em;
}
.aakash-nav-subtitle { font-size: 0.72rem; color: #888; font-weight: 400; }
.aakash-nav-tag {
    background: linear-gradient(135deg, #009FE3 0%, #0078B8 100%);
    color: #fff; font-size: 0.68rem; font-weight: 700;
    letter-spacing: 0.1em; padding: 0.3rem 0.9rem;
    border-radius: 100px; text-transform: uppercase;
}

/* ── Hero strip ── */
.hero-strip {
    background: linear-gradient(135deg, #009FE3 0%, #005F8E 100%);
    padding: 2.2rem 3rem;
    margin-bottom: 1.8rem;
    border-radius: 0 0 20px 20px;
    position: relative; overflow: hidden;
}
.hero-strip::after {
    content:''; position:absolute; top:-60px; right:-60px;
    width:300px; height:300px;
    background: rgba(255,255,255,0.06); border-radius:50%;
}
.hero-eyebrow {
    font-size:0.65rem; font-weight:700; letter-spacing:0.22em;
    color: rgba(255,255,255,0.65); text-transform:uppercase; margin-bottom:0.5rem;
}
.hero-title { font-size:2.4rem; font-weight:800; color:#fff; line-height:1.15; margin-bottom:0.5rem; }
.hero-title .accent { color:#FFE55A; }
.hero-desc { font-size:0.92rem; color:rgba(255,255,255,0.82); max-width:560px; line-height:1.7; }
.badge-row { display:flex; gap:0.6rem; margin-top:1.2rem; flex-wrap:wrap; }
.badge {
    display:inline-flex; align-items:center; gap:0.3rem;
    border-radius:100px; padding:0.26rem 0.85rem;
    font-size:0.69rem; font-weight:700; letter-spacing:0.04em;
}
.badge-yellow { background:rgba(255,229,90,0.2); border:1px solid rgba(255,229,90,0.5); color:#FFE55A; }
.badge-red    { background:rgba(255,255,255,0.12); border:1px solid rgba(255,100,100,0.5); color:#FFAAAA; }
.badge-green  { background:rgba(255,255,255,0.12); border:1px solid rgba(100,230,150,0.5); color:#80EFB0; }
.badge-white  { background:rgba(255,255,255,0.12); border:1px solid rgba(255,255,255,0.3); color:#fff; }

/* ── Section label ── */
.slabel {
    font-size:0.63rem; font-weight:700; letter-spacing:0.2em;
    color:#009FE3; text-transform:uppercase; margin-bottom:0.7rem; margin-top:0.2rem;
}

/* ── Upload cards ── */
.ucard {
    background:#FFFFFF; border:2px dashed #BFD9F5; border-radius:14px;
    padding:1.3rem 1.4rem 0.7rem 1.4rem; transition:border-color 0.2s;
    box-shadow: 0 2px 10px rgba(0,159,227,0.05);
}
.ucard:hover { border-color:#009FE3; }
.utag {
    display:inline-block; font-size:0.64rem; font-weight:700;
    letter-spacing:0.12em; text-transform:uppercase;
    padding:0.22rem 0.75rem; border-radius:100px; margin-bottom:0.7rem;
}
.utag-old { background:rgba(229,37,38,0.08); border:1px solid rgba(229,37,38,0.3); color:#E52526; }
.utag-new { background:rgba(0,159,227,0.1); border:1px solid rgba(0,159,227,0.35); color:#009FE3; }
.ufile-ok { font-size:0.76rem; color:#2ECC71; padding:0.3rem 0 0.7rem 0; font-weight:600; }
.ufile-hint { font-size:0.76rem; color:#AAC4D8; padding:0.3rem 0 0.7rem 0; }

/* ── File uploader ── */
[data-testid="stFileUploader"] > div {
    background:#F7FBFF !important; border:1px solid #D0E8F7 !important; border-radius:10px !important;
}
[data-testid="stFileUploaderDropzoneInstructions"] { color:#AAC4D8 !important; }

/* ── Buttons ── */
.stButton > button {
    background: linear-gradient(135deg, #009FE3 0%, #0078B8 100%) !important;
    color:#fff !important; font-family:'Inter',sans-serif !important;
    font-weight:700 !important; font-size:0.9rem !important; letter-spacing:0.04em !important;
    border:none !important; border-radius:10px !important;
    padding:0.76rem 2rem !important; width:100% !important;
    box-shadow:0 4px 18px rgba(0,159,227,0.35) !important;
    transition:all 0.2s ease !important;
}
.stButton > button:hover { transform:translateY(-1px) !important; box-shadow:0 6px 26px rgba(0,159,227,0.5) !important; }

/* ── Metrics ── */
[data-testid="metric-container"] {
    background:#FFFFFF !important; border:1px solid #E2EFF9 !important;
    border-radius:14px !important; padding:1.2rem 1.5rem !important;
    box-shadow: 0 2px 12px rgba(0,100,180,0.07) !important;
}
[data-testid="stMetricValue"] { font-weight:800 !important; font-size:2rem !important; color:#1A2B3C !important; }
[data-testid="stMetricLabel"] { font-size:0.7rem !important; font-weight:700 !important; color:#009FE3 !important; letter-spacing:0.1em !important; text-transform:uppercase !important; }

/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"] {
    background:#FFFFFF; border-radius:10px; padding:0.3rem;
    border:1px solid #E0EEF8; gap:0.2rem;
    box-shadow: 0 2px 8px rgba(0,100,180,0.05);
}
.stTabs [data-baseweb="tab"] {
    background:transparent !important; color:#6BA8CC !important;
    font-size:0.8rem !important; font-weight:500 !important;
    border-radius:7px !important; padding:0.5rem 1.1rem !important;
    border:none !important;
}
.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, #009FE3 0%, #0078B8 100%) !important;
    color:#fff !important; font-weight:700 !important;
}
.stTabs [data-baseweb="tab-panel"] { padding-top:1rem !important; }

/* ── Dataframe – cleaner look ── */
[data-testid="stDataFrame"] {
    border-radius:12px !important; overflow:hidden !important;
    border:1px solid #D6EAFA !important;
    box-shadow: 0 2px 10px rgba(0,100,180,0.06) !important;
}

/* ── Cards ── */
.card {
    background:#FFFFFF; border:1px solid #E2EFF9; border-radius:14px;
    padding:1.4rem 1.6rem; margin-top:1rem;
    box-shadow: 0 2px 10px rgba(0,100,180,0.06);
}
.card-title { font-size:0.68rem; font-weight:700; color:#009FE3; letter-spacing:0.15em; text-transform:uppercase; margin-bottom:1rem; }

/* ── Download buttons ── */
[data-testid="stDownloadButton"] > button {
    background:#FFFFFF !important; color:#009FE3 !important;
    border:2px solid #BFD9F5 !important; border-radius:10px !important;
    font-family:'Inter',sans-serif !important; font-weight:600 !important; font-size:0.83rem !important;
    width:100% !important; padding:0.7rem 1.2rem !important; transition:all 0.2s !important;
}
[data-testid="stDownloadButton"] > button:hover { background:#F0F8FF !important; border-color:#009FE3 !important; color:#0078B8 !important; }

/* ── Email card ── */
.email-card {
    background: linear-gradient(135deg,#FFFFFF 0%,#EBF6FD 100%);
    border:1.5px solid #BFD9F5; border-radius:16px; padding:1.8rem 2rem; margin-top:1.5rem;
}
.email-title { font-size:1rem; font-weight:700; color:#1A2B3C; margin-bottom:0.3rem; }
.email-sub { font-size:0.82rem; color:#6BA8CC; margin-bottom:1.2rem; }

/* ── Input overrides ── */
[data-testid="stTextInput"] input {
    background:#F7FBFF !important; border:1px solid #D0E8F7 !important;
    color:#1A2B3C !important; border-radius:8px !important; font-size:0.85rem !important;
}
[data-testid="stTextInput"] input:focus { border-color:#009FE3 !important; box-shadow:0 0 0 2px rgba(0,159,227,0.15) !important; }
[data-testid="stTextInput"] label { color:#6BA8CC !important; font-size:0.78rem !important; font-weight:500 !important; }

/* ── Alerts ── */
.stSuccess { background:rgba(46,204,113,0.08) !important; border:1px solid rgba(46,204,113,0.3) !important; color:#1A7A46 !important; border-radius:10px !important; }
.stInfo    { background:rgba(0,159,227,0.07) !important; border:1px solid rgba(0,159,227,0.2) !important; color:#0078B8 !important; border-radius:10px !important; }
.stError   { background:rgba(229,37,38,0.06) !important; border:1px solid rgba(229,37,38,0.2) !important; color:#B71C1C !important; border-radius:10px !important; }

/* ── Divider ── */
.div-line { height:1px; background:linear-gradient(90deg,transparent,#D0E8F7,transparent); margin:1.8rem 0; }

/* ── Change pill colours in tabs ── */
.pill-updated { color:#D68000; font-weight:600; }
.pill-deleted { color:#E52526; font-weight:600; }
.pill-added   { color:#1A7A46; font-weight:600; }

/* ── Summary cards row ── */
.summary-row { display:flex; gap:1rem; flex-wrap:wrap; margin-bottom:1.4rem; }
.scard {
    flex:1; min-width:130px; border-radius:14px; padding:1rem 1.2rem;
    background:#fff; border:1px solid; text-align:center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.04);
}
.scard-num { font-size:2rem; font-weight:800; }
.scard-label { font-size:0.67rem; font-weight:700; letter-spacing:0.12em; text-transform:uppercase; margin-top:0.2rem; }
.scard-upd { border-color:#F5C842; }
.scard-upd .scard-num { color:#D68000; }
.scard-upd .scard-label { color:#D68000; }
.scard-del { border-color:#F5AAAA; }
.scard-del .scard-num { color:#E52526; }
.scard-del .scard-label { color:#E52526; }
.scard-add { border-color:#A5E6C0; }
.scard-add .scard-num { color:#1A7A46; }
.scard-add .scard-label { color:#1A7A46; }
.scard-new { border-color:#C0DFFF; }
.scard-new .scard-num { color:#009FE3; }
.scard-new .scard-label { color:#009FE3; }
.scard-mis { border-color:#E0C0FF; }
.scard-mis .scard-num { color:#7B2DBF; }
.scard-mis .scard-label { color:#7B2DBF; }
</style>
""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════
# Fill colours for Excel
# ═══════════════════════════════════════════════════════════════════
FILL_UPDATED = PatternFill("solid", fgColor="FFF176")
FILL_DELETED = PatternFill("solid", fgColor="FFCDD2")
FILL_ADDED   = PatternFill("solid", fgColor="C8E6C9")
FILL_HEADER  = PatternFill("solid", fgColor="0078B8")
FONT_HEADER  = Font(bold=True, color="FFFFFF", size=10)
FONT_UPDATED = Font(bold=True, color="E65100", size=9)
FONT_DELETED = Font(bold=True, color="B71C1C", size=9)
FONT_ADDED   = Font(bold=True, color="1B5E20", size=9)
thin   = Side(style="thin", color="D0E8F7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

# ═══════════════════════════════════════════════════════════════════
# Core logic
# ═══════════════════════════════════════════════════════════════════
def normalize(val):
    if pd.isna(val): return ""
    return str(val).strip()

def load_sheets(uploaded_file):
    uploaded_file.seek(0)
    return pd.read_excel(uploaded_file, sheet_name=None, header=None, dtype=object)

def compare_workbooks(old_sheets, new_sheets):
    rows = []; cells_map = {}
    old_names = set(old_sheets.keys()); new_names = set(new_sheets.keys())
    for sn in sorted(old_names | new_names):
        if sn not in old_sheets:
            rows.append({"Sheet":sn,"Change Type":"NEW_SHEET","Row":"","Column":"","Old Value":"","New Value":"(entire sheet added)"}); continue
        if sn not in new_sheets:
            rows.append({"Sheet":sn,"Change Type":"MISSING_SHEET","Row":"","Column":"","Old Value":"(entire sheet removed)","New Value":""}); continue
        old_df = old_sheets[sn].copy(); new_df = new_sheets[sn].copy()
        max_cols = max(old_df.shape[1] if old_df.shape[1] else 0, new_df.shape[1] if new_df.shape[1] else 0)
        old_df = old_df.reindex(columns=range(max_cols)); new_df = new_df.reindex(columns=range(max_cols))
        old_rc = len(old_df); new_rc = len(new_df); max_rows = max(old_rc, new_rc)
        cells_map.setdefault(sn, [])
        for r in range(max_rows):
            excel_row = r + 1
            if r >= new_rc:
                old_row_vals = [normalize(old_df.iat[r, c]) for c in range(max_cols)]
                if any(v != "" for v in old_row_vals):
                    rows.append({"Sheet":sn,"Change Type":"DELETED_ROW","Row":excel_row,"Column":"All","Old Value":" | ".join(old_row_vals),"New Value":"(row deleted)"})
                    for c in range(max_cols): cells_map[sn].append((excel_row, c+1, "DELETED"))
                continue
            if r >= old_rc:
                new_row_vals = [normalize(new_df.iat[r, c]) for c in range(max_cols)]
                if any(v != "" for v in new_row_vals):
                    rows.append({"Sheet":sn,"Change Type":"ADDED_ROW","Row":excel_row,"Column":"All","Old Value":"(row added)","New Value":" | ".join(new_row_vals)})
                    for c in range(max_cols): cells_map[sn].append((excel_row, c+1, "ADDED"))
                continue
            for c in range(max_cols):
                ov = normalize(old_df.iat[r, c]); nv = normalize(new_df.iat[r, c])
                if ov != nv:
                    rows.append({"Sheet":sn,"Change Type":"UPDATED_CELL","Row":excel_row,"Column":c+1,"Old Value":ov,"New Value":nv})
                    cells_map[sn].append((excel_row, c+1, "UPDATED"))
    return (pd.DataFrame(rows) if rows else pd.DataFrame(columns=["Sheet","Change Type","Row","Column","Old Value","New Value"])), cells_map

def highlight_new_file(new_file, cells_map):
    new_file.seek(0)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(new_file.read()); tmp_path = tmp.name
    wb = load_workbook(tmp_path)
    fill_map = {"UPDATED": FILL_UPDATED, "DELETED": FILL_DELETED, "ADDED": FILL_ADDED}
    for sn, cell_list in cells_map.items():
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        for r, c, ctype in cell_list:
            ws.cell(row=r, column=c).fill = fill_map.get(ctype, FILL_UPDATED)
    out = BytesIO(); wb.save(out); out.seek(0); wb.close(); os.remove(tmp_path)
    return out

def build_master_sheet(report_df, highlighted_file_bytes):
    wb = Workbook()
    ws_sum = wb.active; ws_sum.title = "Summary"
    ws_sum.merge_cells("A1:F1"); ws_sum["A1"] = "Aakash ExcelDiff Pro — Master Change Report"
    ws_sum["A1"].font = Font(bold=True, size=14, color="FFFFFF"); ws_sum["A1"].fill = FILL_HEADER
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center"); ws_sum.row_dimensions[1].height = 32
    ws_sum.merge_cells("A2:F2"); ws_sum["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M:%S')}"
    ws_sum["A2"].font = Font(italic=True, size=10, color="5A8AB0"); ws_sum["A2"].alignment = Alignment(horizontal="center"); ws_sum.row_dimensions[2].height = 20
    ws_sum["A4"] = "Change Type"; ws_sum["B4"] = "Count"
    for cell in [ws_sum["A4"], ws_sum["B4"]]:
        cell.font = FONT_HEADER; cell.fill = FILL_HEADER; cell.alignment = Alignment(horizontal="center")
    counts = {
        "Updated Cells":  len(report_df[report_df["Change Type"]=="UPDATED_CELL"])  if not report_df.empty else 0,
        "Deleted Rows":   len(report_df[report_df["Change Type"]=="DELETED_ROW"])   if not report_df.empty else 0,
        "Added Rows":     len(report_df[report_df["Change Type"]=="ADDED_ROW"])     if not report_df.empty else 0,
        "New Sheets":     len(report_df[report_df["Change Type"]=="NEW_SHEET"])     if not report_df.empty else 0,
        "Missing Sheets": len(report_df[report_df["Change Type"]=="MISSING_SHEET"]) if not report_df.empty else 0,
    }
    fills_summary = {"Updated Cells":FILL_UPDATED,"Deleted Rows":FILL_DELETED,"Added Rows":FILL_ADDED}
    for i,(k,v) in enumerate(counts.items(), start=5):
        ws_sum[f"A{i}"] = k; ws_sum[f"B{i}"] = v
        ws_sum[f"A{i}"].fill = fills_summary.get(k, PatternFill("solid", fgColor="E3F2FD"))
        ws_sum[f"B{i}"].alignment = Alignment(horizontal="center")
    ws_sum.column_dimensions["A"].width = 22; ws_sum.column_dimensions["B"].width = 12

    ws_log = wb.create_sheet("Full Change Log")
    if not report_df.empty:
        headers = list(report_df.columns)
        for ci, h in enumerate(headers, 1):
            cell = ws_log.cell(row=1, column=ci, value=h)
            cell.font = FONT_HEADER; cell.fill = FILL_HEADER; cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
        type_fill_map = {
            "UPDATED_CELL":(FILL_UPDATED,FONT_UPDATED),"DELETED_ROW":(FILL_DELETED,FONT_DELETED),
            "ADDED_ROW":(FILL_ADDED,FONT_ADDED),
            "NEW_SHEET":(PatternFill("solid",fgColor="EDE7F6"),Font(bold=True,color="4A148C",size=9)),
            "MISSING_SHEET":(PatternFill("solid",fgColor="FCE4EC"),Font(bold=True,color="880E4F",size=9)),
        }
        for ri, row_data in enumerate(report_df.itertuples(index=False), start=2):
            ctype = row_data[1]; fill, font = type_fill_map.get(ctype, (None,None))
            for ci, val in enumerate(row_data, start=1):
                c = ws_log.cell(row=ri, column=ci, value=val)
                if ci == 2 and fill: c.fill = fill; c.font = font
                c.border = BORDER; c.alignment = Alignment(wrap_text=True, vertical="top")
        for ci in range(1, len(headers)+1):
            ws_log.column_dimensions[get_column_letter(ci)].width = 20

    if not report_df.empty:
        for sn in report_df["Sheet"].dropna().unique():
            safe_name = (str(sn)[:28] + "…") if len(str(sn)) > 28 else str(sn)
            ws_s = wb.create_sheet(f"Sheet-{safe_name}"[:31])
            sub = report_df[report_df["Sheet"]==sn]
            ws_s.merge_cells("A1:F1"); ws_s["A1"] = f"Sheet: {sn}  —  {len(sub)} change(s)"
            ws_s["A1"].font = Font(bold=True,size=11,color="FFFFFF"); ws_s["A1"].fill = FILL_HEADER
            ws_s["A1"].alignment = Alignment(horizontal="center",vertical="center"); ws_s.row_dimensions[1].height = 26
            headers = list(report_df.columns)
            for ci, h in enumerate(headers, 1):
                cell = ws_s.cell(row=2, column=ci, value=h)
                cell.font = FONT_HEADER; cell.fill = PatternFill("solid",fgColor="0078B8")
                cell.alignment = Alignment(horizontal="center"); cell.border = BORDER
            type_fill_map2 = {"UPDATED_CELL":FILL_UPDATED,"DELETED_ROW":FILL_DELETED,"ADDED_ROW":FILL_ADDED}
            for ri, row_data in enumerate(sub.itertuples(index=False), start=3):
                ctype = row_data[1]; fill2 = type_fill_map2.get(ctype)
                for ci, val in enumerate(row_data, start=1):
                    c = ws_s.cell(row=ri, column=ci, value=val)
                    if ci == 2 and fill2: c.fill = fill2
                    c.border = BORDER; c.alignment = Alignment(wrap_text=True)
            for ci in range(1, len(headers)+1):
                ws_s.column_dimensions[get_column_letter(ci)].width = 20

    out = BytesIO(); wb.save(out); out.seek(0)
    return out

def send_email_report(sender_email, sender_password, report_bytes, highlighted_bytes, recipient="pankajshukla@aesl.in"):
    msg = MIMEMultipart()
    msg["From"] = sender_email; msg["To"] = recipient
    msg["Subject"] = f"Aakash ExcelDiff Pro — Change Report [{datetime.now().strftime('%d %b %Y')}]"
    body = f"""Dear Pankaj,

Please find attached the Excel comparison report generated by Aakash ExcelDiff Pro.

📊 Master Report file contains the full summary.
🔆 Highlighted file shows all changes colour-coded:
   🟡 Yellow  = Updated cells
   🔴 Red     = Deleted rows
   🟢 Green   = Added rows

Generated on {datetime.now().strftime('%d %b %Y at %H:%M')}.

Regards,
Aakash ExcelDiff Pro""".strip()
    msg.attach(MIMEText(body, "plain"))
    for fname, fbytes in [("Master_Change_Report.xlsx", report_bytes), ("Highlighted_New_File.xlsx", highlighted_bytes)]:
        part = MIMEBase("application", "octet-stream"); part.set_payload(fbytes.read())
        encoders.encode_base64(part); part.add_header("Content-Disposition", f"attachment; filename={fname}")
        msg.attach(part); fbytes.seek(0)
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender_email, sender_password); server.sendmail(sender_email, recipient, msg.as_string())

# ───────────────────────────────────────────────────────────────────
# Helper: styled dataframe with colour-coded Change Type column
# ───────────────────────────────────────────────────────────────────
def colour_change_type(val):
    colour_map = {
        "UPDATED_CELL":  "background-color:#FFF9C4; color:#D68000; font-weight:700;",
        "DELETED_ROW":   "background-color:#FFEBEE; color:#E52526; font-weight:700;",
        "ADDED_ROW":     "background-color:#E8F5E9; color:#1A7A46; font-weight:700;",
        "NEW_SHEET":     "background-color:#EDE7F6; color:#6A1B9A; font-weight:700;",
        "MISSING_SHEET": "background-color:#FCE4EC; color:#880E4F; font-weight:700;",
    }
    return colour_map.get(val, "")

def show_styled_table(df):
    if df.empty:
        st.info("No changes in this category.")
        return
    styled = df.style.applymap(colour_change_type, subset=["Change Type"])
    st.dataframe(styled, use_container_width=True, height=420)

# ═══════════════════════════════════════════════════════════════════
# UI
# ═══════════════════════════════════════════════════════════════════

# ── Nav bar ──
st.markdown(f"""
<div class="aakash-nav">
  <div class="aakash-nav-logo">
    <img src="data:image/png;base64,{LOGO_B64}" alt="Aakash Logo"/>
    <div>
     
   
  </div>
  <div class="aakash-nav-tag">Internal Tool</div>
</div>
""", unsafe_allow_html=True)

# ── Hero ──
st.markdown("""
<div class="hero-strip">
  <div class="hero-eyebrow">⚡ Spreadsheet Intelligence</div>
  <div class="hero-title">Excel<span class="accent">Diff</span> Pro</div>
  <div class="hero-desc">Cell-by-cell comparison across every sheet — instantly detects updated, deleted, and added rows. Download a master report or email it in one click.</div>
  <div class="badge-row">
    <span class="badge badge-yellow">🟡 Updated Cells</span>
    <span class="badge badge-red">🔴 Deleted Rows</span>
    <span class="badge badge-green">🟢 Added Rows</span>
    <span class="badge badge-white">📄 Multi-sheet</span>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Upload ──
st.markdown('<div class="slabel">Step 1 — Upload files</div>', unsafe_allow_html=True)
col1, col2 = st.columns(2, gap="large")

with col1:
    st.markdown('<div class="ucard"><span class="utag utag-old">◈ Baseline (Old)</span>', unsafe_allow_html=True)
    old_file = st.file_uploader("old", type=["xlsx"], key="old", label_visibility="collapsed")
    if old_file:
        st.markdown(f'<div class="ufile-ok">✓ {old_file.name}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="ufile-hint">Drop original .xlsx here</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    st.markdown('<div class="ucard"><span class="utag utag-new">◈ Updated (New)</span>', unsafe_allow_html=True)
    new_file = st.file_uploader("new", type=["xlsx"], key="new", label_visibility="collapsed")
    if new_file:
        st.markdown(f'<div class="ufile-ok">✓ {new_file.name}</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="ufile-hint">Drop revised .xlsx here</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

st.markdown('<div style="height:1.2rem"></div>', unsafe_allow_html=True)

if old_file and new_file:
    st.markdown('<div class="slabel">Step 2 — Run Comparison</div>', unsafe_allow_html=True)
    cb, _ = st.columns([1,2])
    with cb:
        run = st.button("⚡  Compare Files", use_container_width=True)

    if run:
        with st.spinner("Analysing sheets…"):
            old_sheets = load_sheets(old_file)
            new_sheets = load_sheets(new_file)
            report_df, cells_map = compare_workbooks(old_sheets, new_sheets)
            highlighted_bytes = highlight_new_file(new_file, cells_map)
            master_bytes = build_master_sheet(report_df, highlighted_bytes)
            highlighted_bytes.seek(0)
        st.session_state["report_df"]         = report_df
        st.session_state["highlighted_bytes"] = highlighted_bytes
        st.session_state["master_bytes"]      = master_bytes

    if "report_df" in st.session_state:
        report_df         = st.session_state["report_df"]
        highlighted_bytes = st.session_state["highlighted_bytes"]
        master_bytes      = st.session_state["master_bytes"]

        st.success("✓ Comparison complete!")
        st.markdown('<div style="height:0.8rem"></div>', unsafe_allow_html=True)

        n_upd = len(report_df[report_df["Change Type"]=="UPDATED_CELL"])  if not report_df.empty else 0
        n_del = len(report_df[report_df["Change Type"]=="DELETED_ROW"])   if not report_df.empty else 0
        n_add = len(report_df[report_df["Change Type"]=="ADDED_ROW"])     if not report_df.empty else 0
        n_ns  = len(report_df[report_df["Change Type"]=="NEW_SHEET"])     if not report_df.empty else 0
        n_ms  = len(report_df[report_df["Change Type"]=="MISSING_SHEET"]) if not report_df.empty else 0

        # ── Summary cards ──
        st.markdown('<div class="slabel">Summary</div>', unsafe_allow_html=True)
        st.markdown(f"""
        <div class="summary-row">
          <div class="scard scard-upd"><div class="scard-num">{n_upd}</div><div class="scard-label">🟡 Updated Cells</div></div>
          <div class="scard scard-del"><div class="scard-num">{n_del}</div><div class="scard-label">🔴 Deleted Rows</div></div>
          <div class="scard scard-add"><div class="scard-num">{n_add}</div><div class="scard-label">🟢 Added Rows</div></div>
          <div class="scard scard-new"><div class="scard-num">{n_ns}</div><div class="scard-label">📄 New Sheets</div></div>
          <div class="scard scard-mis"><div class="scard-num">{n_ms}</div><div class="scard-label">⚠️ Missing Sheets</div></div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="div-line"></div>', unsafe_allow_html=True)

        # ── Tabs with styled tables ──
        st.markdown('<div class="slabel">Detailed Change Log</div>', unsafe_allow_html=True)
        tab_all, tab_upd, tab_del, tab_add, tab_sheets = st.tabs([
            f"📋 All ({len(report_df)})",
            f"🟡 Updated ({n_upd})",
            f"🔴 Deleted ({n_del})",
            f"🟢 Added ({n_add})",
            f"📄 Sheets ({n_ns+n_ms})",
        ])
        with tab_all:    show_styled_table(report_df)
        with tab_upd:    show_styled_table(report_df[report_df["Change Type"]=="UPDATED_CELL"])
        with tab_del:    show_styled_table(report_df[report_df["Change Type"]=="DELETED_ROW"])
        with tab_add:    show_styled_table(report_df[report_df["Change Type"]=="ADDED_ROW"])
        with tab_sheets: show_styled_table(report_df[report_df["Change Type"].isin(["NEW_SHEET","MISSING_SHEET"])])

        st.markdown('<div class="div-line"></div>', unsafe_allow_html=True)

        # ── Downloads ──
        st.markdown('<div class="card"><div class="card-title">⬇ Download Files</div>', unsafe_allow_html=True)
        d1, d2 = st.columns(2, gap="medium")
        with d1:
            master_bytes.seek(0)
            st.download_button("📊  Master Report (all sheets)", data=master_bytes.read(),
                file_name="Master_Change_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with d2:
            highlighted_bytes.seek(0)
            st.download_button("🔆  Highlighted New File", data=highlighted_bytes.read(),
                file_name="Highlighted_New_File.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

        # ── Email ──
        st.markdown("""
        <div class="email-card">
          <div class="email-title">📧 Email Report</div>
          <div class="email-sub">Send both files automatically to pankajshukla@aesl.in</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown('<div style="height:0.8rem"></div>', unsafe_allow_html=True)

        with st.expander("⚙️  Configure & Send Email", expanded=False):
            st.markdown('<div style="font-size:0.78rem;color:#6BA8CC;margin-bottom:0.8rem;">Gmail account use karein. 2FA ke saath App Password recommended hai.</div>', unsafe_allow_html=True)
            e1, e2 = st.columns(2, gap="medium")
            with e1: sender_email = st.text_input("Your Gmail Address", placeholder="yourname@gmail.com", key="s_email")
            with e2: sender_pass  = st.text_input("App Password", type="password", placeholder="xxxx xxxx xxxx xxxx", key="s_pass")
            st.markdown('<div style="font-size:0.78rem;color:#1A7A46;margin:0.4rem 0 0.8rem 0;">📬 Recipient: <strong>pankajshukla@aesl.in</strong></div>', unsafe_allow_html=True)
            send_col, _ = st.columns([1,2])
            with send_col:
                send_clicked = st.button("📤  Send Email Now", use_container_width=True, key="send_btn")
            if send_clicked:
                if not sender_email or not sender_pass:
                    st.error("Please enter Gmail address and App Password.")
                else:
                    try:
                        with st.spinner("Sending…"):
                            master_bytes.seek(0); highlighted_bytes.seek(0)
                            send_email_report(sender_email, sender_pass, master_bytes, highlighted_bytes)
                        st.success("✓ Email sent to pankajshukla@aesl.in!")
                    except smtplib.SMTPAuthenticationError:
                        st.error("Authentication failed. App Password check karein.")
                    except Exception as ex:
                        st.error(f"Send failed: {ex}")
            st.markdown("""
            <div style="background:#F7FBFF;border:1px solid #D0E8F7;border-radius:10px;padding:1rem 1.2rem;margin-top:1rem;font-size:0.77rem;color:#6BA8CC;line-height:1.7;">
            <strong style="color:#009FE3;">App Password kaise banayein?</strong><br>
            1. Google Account → Security → 2-Step Verification ON<br>
            2. Security → App Passwords → "Mail" → Generate<br>
            3. Woh 16-character password yahan paste karein
            </div>""", unsafe_allow_html=True)

else:
    st.info("Upload both files - Old (baseline) and New (updated) - then click the Compare button.")
